"""Importación de cuadros de carga desde Excel / CSV / JSON.

Cuando el ingeniero ya tiene un cuadro de carga elaborado (típico en Chile
en formato Excel), puede subirlo y la app extrae los recintos y/o las cargas
dedicadas automáticamente.

Formato esperado (ver plantillas en docs/templates/):

Cuadro de RECINTOS:
    columnas:  nombre  |  uso  |  area_m2  |  perimetro_m (opcional)
    ejemplo:   Living  |  living  |  22.5  |  19

Cuadro de CARGAS DEDICADAS:
    columnas:  nombre  |  potencia_w  |  recinto  |  activa (opcional, default Sí)
    ejemplo:   Cocina/horno eléctrico  |  5500  |  Cocina  |  Sí

El sistema acepta también nombres de columnas en español/inglés y con/sin
acentos, mayúsculas/minúsculas, espacios.
"""
from __future__ import annotations

import csv
import io
import json
import unicodedata
from typing import Any, Iterable

from app.services.usos_recinto import inferir_uso


# Mapeo de aliases de columnas → campos estándar
ALIASES_RECINTO = {
    "nombre":   {"nombre", "recinto", "name", "habitacion", "habitación", "dependencia", "descripcion", "descripción"},
    "uso":      {"uso", "tipo", "use", "destino", "function"},
    "area_m2":  {"area_m2", "area", "área", "m2", "m²", "superficie", "sup_m2", "area_m²", "metros"},
    "perimetro_m": {"perimetro_m", "perímetro_m", "perimetro", "perimeter"},
}

ALIASES_CARGA = {
    "nombre":      {"nombre", "carga", "equipo", "name", "descripcion", "descripción"},
    "potencia_w":  {"potencia_w", "potencia", "watts", "w", "pot_w", "power", "kw"},
    "recinto":     {"recinto", "dependencia", "ubicacion", "ubicación", "location", "sala"},
    "activa":      {"activa", "active", "estado", "habilitada", "on", "uso"},
}


def _norm(s: Any) -> str:
    """Normaliza una clave: lower, sin tildes, quita espacios/underscores/guiones."""
    if s is None: return ""
    s = str(s).strip().lower()
    nfkd = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    # Quitar separadores comunes para hacer match flexible
    return s.replace(" ", "").replace("_", "").replace("-", "")


def _match_alias(header: str, aliases_map: dict[str, set[str]]) -> str | None:
    """Devuelve el campo estándar al que pertenece este header, o None."""
    h = _norm(header)
    for campo, aliases in aliases_map.items():
        if h in {_norm(a) for a in aliases}:
            return campo
    return None


def _parse_bool(v: Any, default: bool = True) -> bool:
    if v is None or v == "": return default
    s = _norm(v)
    if s in {"si", "sí", "yes", "y", "true", "t", "1", "activa", "activo", "on"}: return True
    if s in {"no", "n", "false", "f", "0", "inactiva", "inactivo", "off", "-"}: return False
    return default


def _parse_float(v: Any) -> float | None:
    """Parsea un float aceptando formatos: '3.0', '3,5', '1.234,56', '1,234.56'."""
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(" ", "")
    if not s: return None
    has_comma = "," in s
    has_dot = "." in s
    try:
        if has_comma and has_dot:
            # Determinar cuál es el separador decimal por posición (el último)
            if s.rfind(",") > s.rfind("."):
                # Formato chileno: 1.234,56 → quita puntos, cambia coma por punto
                s = s.replace(".", "").replace(",", ".")
            else:
                # Formato anglosajón: 1,234.56 → quita comas
                s = s.replace(",", "")
        elif has_comma:
            # Solo coma: asumir que es decimal (común en CL)
            s = s.replace(",", ".")
        # Si solo tiene punto, asumir que es decimal anglosajón → no tocar
        return float(s)
    except (TypeError, ValueError):
        return None


def _mapear_columnas(headers: list[str], aliases_map: dict[str, set[str]]) -> dict[str, int]:
    """Devuelve {campo_estandar: idx_columna}. Si una columna no matchea ningún alias, se ignora."""
    out: dict[str, int] = {}
    for idx, h in enumerate(headers):
        campo = _match_alias(h, aliases_map)
        if campo and campo not in out:
            out[campo] = idx
    return out


def parse_csv_recintos(contenido_csv: str) -> list[dict]:
    """Parsea CSV y retorna lista de recintos validados."""
    reader = csv.reader(io.StringIO(contenido_csv))
    rows = list(reader)
    if not rows: raise ValueError("CSV vacío")
    # Detectar separador automático (",", ";", "\t")
    sample = contenido_csv[:1000]
    sep = max([",", ";", "\t"], key=lambda c: sample.count(c))
    rows = list(csv.reader(io.StringIO(contenido_csv), delimiter=sep))
    headers, *data_rows = rows
    mapa = _mapear_columnas(headers, ALIASES_RECINTO)
    if "nombre" not in mapa or "area_m2" not in mapa:
        raise ValueError(f"Columnas mínimas no encontradas (nombre + area_m2). Headers leídos: {headers}")
    return _filas_a_recintos(data_rows, mapa)


def _filas_a_recintos(data_rows: Iterable, mapa: dict[str, int]) -> list[dict]:
    out = []
    for i, row in enumerate(data_rows, start=1):
        if not any(cell for cell in row): continue  # fila vacía
        nombre = str(row[mapa["nombre"]]).strip() if mapa["nombre"] < len(row) else ""
        if not nombre: continue
        area = _parse_float(row[mapa["area_m2"]]) if mapa["area_m2"] < len(row) else None
        if not area or area <= 0: continue
        uso_raw = str(row[mapa["uso"]]).strip() if "uso" in mapa and mapa["uso"] < len(row) else ""
        uso = inferir_uso(uso_raw) if uso_raw else inferir_uso(nombre)
        perim = _parse_float(row[mapa["perimetro_m"]]) if "perimetro_m" in mapa and mapa["perimetro_m"] < len(row) else 0
        out.append({
            "id": i, "nombre": nombre, "uso": uso, "area_m2": round(area, 2),
            "perimetro_m": round(perim or 0, 1),
        })
    return out


def parse_csv_cargas(contenido_csv: str) -> list[dict]:
    """Parsea CSV de cargas dedicadas."""
    sample = contenido_csv[:1000]
    sep = max([",", ";", "\t"], key=lambda c: sample.count(c))
    rows = list(csv.reader(io.StringIO(contenido_csv), delimiter=sep))
    if not rows: raise ValueError("CSV vacío")
    headers, *data_rows = rows
    mapa = _mapear_columnas(headers, ALIASES_CARGA)
    if "nombre" not in mapa or "potencia_w" not in mapa:
        raise ValueError(f"Columnas mínimas no encontradas (nombre + potencia_w). Headers: {headers}")
    out = []
    for row in data_rows:
        if not any(cell for cell in row): continue
        nombre = str(row[mapa["nombre"]]).strip() if mapa["nombre"] < len(row) else ""
        if not nombre: continue
        pot_raw = row[mapa["potencia_w"]] if mapa["potencia_w"] < len(row) else None
        # Detectar si la columna estaba en kW (header "kw")
        pot = _parse_float(pot_raw)
        if pot is None or pot <= 0: continue
        # Si el header original era kW, convertir
        header_potencia = _norm(headers[mapa["potencia_w"]])
        if header_potencia in {"kw", "kilowatts", "kw_max"}:
            pot = pot * 1000
        recinto = str(row[mapa["recinto"]]).strip() if "recinto" in mapa and mapa["recinto"] < len(row) else ""
        activa = _parse_bool(row[mapa["activa"]] if "activa" in mapa and mapa["activa"] < len(row) else None)
        out.append({
            "nombre": nombre,
            "potencia_w": round(pot, 0),
            "recinto": recinto or "__general__",
            "activa": activa,
        })
    return out


def parse_xlsx_bytes(xlsx_bytes: bytes, tipo: str) -> list[dict]:
    """Lee un Excel binario con openpyxl y devuelve la lista parseada."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    # Tomar la primera hoja
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: raise ValueError("Excel vacío")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = [list(r) for r in rows[1:]]
    if tipo == "recintos":
        mapa = _mapear_columnas(headers, ALIASES_RECINTO)
        if "nombre" not in mapa or "area_m2" not in mapa:
            raise ValueError(f"Columnas no encontradas (nombre + area_m2). Headers: {headers}")
        return _filas_a_recintos(data_rows, mapa)
    elif tipo == "cargas":
        mapa = _mapear_columnas(headers, ALIASES_CARGA)
        if "nombre" not in mapa or "potencia_w" not in mapa:
            raise ValueError(f"Columnas no encontradas. Headers: {headers}")
        out = []
        for row in data_rows:
            if not any(cell is not None and cell != "" for cell in row): continue
            nombre = str(row[mapa["nombre"]]).strip() if row[mapa["nombre"]] else ""
            if not nombre: continue
            pot = _parse_float(row[mapa["potencia_w"]] if mapa["potencia_w"] < len(row) else None)
            if pot is None: continue
            header_pot = _norm(headers[mapa["potencia_w"]])
            if header_pot in {"kw", "kilowatts"}: pot *= 1000
            recinto = str(row[mapa["recinto"]]).strip() if "recinto" in mapa and row[mapa["recinto"]] else ""
            activa = _parse_bool(row[mapa["activa"]] if "activa" in mapa and mapa["activa"] < len(row) else None)
            out.append({
                "nombre": nombre, "potencia_w": round(pot, 0),
                "recinto": recinto or "__general__", "activa": activa,
            })
        return out
    raise ValueError("tipo debe ser 'recintos' o 'cargas'")


def plantilla_csv_recintos() -> str:
    return (
        "nombre,uso,area_m2,perimetro_m\n"
        "Living comedor,living,22.5,19\n"
        "Cocina,cocina,10.5,13\n"
        "Baño 1,bano,3.0,7\n"
        "Dormitorio 1,dormitorio,13.5,15\n"
        "Dormitorio 2,dormitorio,12.0,14\n"
    )


def plantilla_csv_cargas() -> str:
    return (
        "nombre,potencia_w,recinto,activa\n"
        "Cocina eléctrica,5500,Cocina,Sí\n"
        "Microondas,1500,Cocina,Sí\n"
        "Lavadora,2200,Cocina,Sí\n"
        "Calefón eléctrico,5500,Baño 1,No\n"
        "Aire acondicionado dormitorio,2200,Dormitorio 1,Sí\n"
    )
