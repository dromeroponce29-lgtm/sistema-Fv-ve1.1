"""Generador de plantilla Excel profesional para cuadros de carga.

Esta plantilla la entrega el ingeniero a su cliente para que llene los
recintos y las cargas de su proyecto. Cuando la devuelve, el ingeniero la
sube por la UI (paso 3 y 4 del wizard) y la app importa automáticamente
todos los datos.

Hojas:
  1. Instrucciones — cómo llenar la planilla
  2. Recintos     — dropdown de uso, ejemplo pre-cargado
  3. Cargas       — dropdown enlazado a recintos de la hoja 2
"""
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

ORANGE = "C56022"
ORANGE_DARK = "8C3F12"
ORANGE_LIGHT = "F5E6D8"
ORANGE_BG = "FAF1E6"
GRAY = "595959"
GREEN_OK = "3A8348"

USOS_RIC = [
    "living", "comedor", "dormitorio", "cocina", "bano", "oficina",
    "pasillo", "hall", "exterior", "lavanderia", "logia", "bodega",
    "circulacion", "comun", "desconocido",
]


def _set_border(cell, all_sides=True, color="BFBFBF"):
    side = Side(style="thin", color=color)
    if all_sides:
        cell.border = Border(top=side, bottom=side, left=side, right=side)


def _header_row(ws, row, headers, col_widths=None):
    """Estampa la fila de header con estilo."""
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ORANGE_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _set_border(c)
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    ws.row_dimensions[row].height = 28


def _data_row(ws, row, values, formats=None, alt_row=False):
    """Estampa una fila de datos."""
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="Calibri", size=10.5)
        c.alignment = Alignment(horizontal="left" if isinstance(v, str) else "right", vertical="center")
        if alt_row:
            c.fill = PatternFill("solid", fgColor=ORANGE_BG)
        if formats and i <= len(formats) and formats[i-1]:
            c.number_format = formats[i-1]
        _set_border(c)


def _title(ws, row, col, text, span_cols=1, color=ORANGE_DARK, size=14):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=size, bold=True, color=color)
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span_cols - 1)


def generar_template_xlsx() -> bytes:
    """Crea y devuelve el .xlsx como bytes (para entregar por HTTP)."""
    wb = Workbook()

    # =====================================================================
    # HOJA 1: INSTRUCCIONES
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Instrucciones"
    ws1.sheet_view.showGridLines = False

    _title(ws1, 1, 1, "CUADRO DE CARGA — SISTEMAS FOTOVOLTAICOS CHILE", span_cols=6, size=16)
    ws1.cell(row=2, column=1, value="Plantilla profesional · v1.0 · llena las hojas siguientes y sube este archivo a la app FV Chile").font = Font(name="Calibri", size=10, italic=True, color=GRAY)

    # Marco con instrucciones
    instrucciones = [
        ("CÓMO USAR ESTA PLANTILLA", 14, True, ORANGE_DARK),
        ("", 10, False, GRAY),
        ("Esta plantilla tiene dos hojas que debes llenar:", 11, False, "1F1D18"),
        ("", 10, False, GRAY),
        ("Hoja \"Recintos\"", 12, True, ORANGE_DARK),
        ("   • Una fila por cada habitación o dependencia del proyecto.", 11, False, "1F1D18"),
        ("   • Columnas obligatorias: nombre · uso (dropdown) · área en m².", 11, False, "1F1D18"),
        ("   • La columna \"uso\" tiene un dropdown con los valores válidos para el cálculo RIC.", 11, False, "1F1D18"),
        ("   • Puedes agregar tantas filas como necesites.", 11, False, "1F1D18"),
        ("", 10, False, GRAY),
        ("Hoja \"Cargas\"", 12, True, ORANGE_DARK),
        ("   • Una fila por cada equipo eléctrico dedicado (cocina, calefón, aire, motores, etc).", 11, False, "1F1D18"),
        ("   • Columnas: nombre · potencia (W) · recinto donde se ubica · activa (Sí/No).", 11, False, "1F1D18"),
        ("   • La columna \"recinto\" tiene un dropdown que se llena con los nombres que pusiste en la hoja anterior.", 11, False, "1F1D18"),
        ("", 10, False, GRAY),
        ("Una vez llena", 12, True, GREEN_OK),
        ("   1. Guarda este archivo (Cmd+S / Ctrl+S).", 11, False, "1F1D18"),
        ("   2. Abre la app FV Chile → \"+ Nuevo proyecto\".", 11, False, "1F1D18"),
        ("   3. En paso 3 (Recintos) y paso 4 (Cargas), súbelo con el botón \"Seleccionar archivo\".", 11, False, "1F1D18"),
        ("   4. La app valida los datos, aplica las tablas RIC y avanza con el dimensionamiento FV.", 11, False, "1F1D18"),
        ("", 10, False, GRAY),
        ("Aliases aceptados en los headers (no se distingue mayúsculas/acentos/espacios):", 11, True, GRAY),
        ("   nombre = recinto / habitación / dependencia / descripción", 10, False, GRAY),
        ("   area_m2 = área / superficie / m² / metros", 10, False, GRAY),
        ("   uso = tipo / destino", 10, False, GRAY),
        ("   potencia_w = potencia / watts / W (si la columna se llama \"kW\", se multiplica por 1.000)", 10, False, GRAY),
        ("   recinto = dependencia / ubicación / sala", 10, False, GRAY),
        ("", 10, False, GRAY),
        ("Soporte: dromeroponce29@gmail.com · FV Chile v1.0", 9, False, GRAY),
    ]
    for i, (text, size, bold, color) in enumerate(instrucciones, start=4):
        c = ws1.cell(row=i, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    ws1.column_dimensions['A'].width = 8
    for col in 'BCDEFGH': ws1.column_dimensions[col].width = 12

    # =====================================================================
    # HOJA 2: RECINTOS
    # =====================================================================
    ws2 = wb.create_sheet("Recintos")
    ws2.sheet_view.showGridLines = False
    _title(ws2, 1, 1, "RECINTOS DEL PROYECTO", span_cols=5, size=14)
    ws2.cell(row=2, column=1, value="Lista las habitaciones/dependencias del proyecto. La columna \"uso\" tiene dropdown.").font = Font(name="Calibri", size=10, italic=True, color=GRAY)
    ws2.merge_cells("A2:E2")

    headers = ["Nombre", "Uso (dropdown)", "Área (m²)", "Perímetro (m, opcional)", "Notas"]
    _header_row(ws2, 4, headers, col_widths=[28, 18, 13, 20, 30])

    # Ejemplos pre-cargados (vivienda 2D-1B)
    ejemplos = [
        ("Living comedor", "living",     22.5, 19, "Espacio principal con luz natural"),
        ("Cocina",         "cocina",     10.5, 13, "Incluye barra y comedor diario"),
        ("Baño",           "bano",        3.0,  7, ""),
        ("Dormitorio 1",   "dormitorio", 13.5, 15, "Habitación principal con clóset"),
        ("Dormitorio 2",   "dormitorio", 12.0, 14, ""),
    ]
    for i, ej in enumerate(ejemplos):
        _data_row(ws2, 5 + i, ej, formats=[None, None, "0.00", "0.0", None], alt_row=(i % 2 == 1))

    # Fila de filas vacías para que el usuario las llene
    for i in range(len(ejemplos), len(ejemplos) + 15):
        _data_row(ws2, 5 + i, ["", "", "", "", ""], formats=[None, None, "0.00", "0.0", None], alt_row=(i % 2 == 1))

    # Fila TOTAL con fórmula
    total_row = 5 + len(ejemplos) + 15 + 1
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11, color=ORANGE_DARK)
    ws2.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})").font = Font(bold=True, size=11, color=ORANGE_DARK)
    ws2.cell(row=total_row, column=3).number_format = "0.00\" m²\""
    for col in range(1, 6):
        c = ws2.cell(row=total_row, column=col)
        c.fill = PatternFill("solid", fgColor=ORANGE_LIGHT)
        _set_border(c)

    # Data validation para la columna "uso" (dropdown)
    dv_uso = DataValidation(
        type="list",
        formula1='"' + ",".join(USOS_RIC) + '"',
        allow_blank=True,
        showDropDown=False,  # OJO: False muestra el dropdown (es invertido en xlsx)
    )
    dv_uso.error = "Uso no válido. Selecciona uno del dropdown."
    dv_uso.errorTitle = "Uso inválido"
    dv_uso.prompt = "Elige el uso del recinto (define la potencia W/m² según RIC)"
    dv_uso.promptTitle = "Uso del recinto"
    ws2.add_data_validation(dv_uso)
    dv_uso.add(f"B5:B{total_row-1}")

    # Data validation para área (debe ser positivo)
    dv_area = DataValidation(type="decimal", operator="greaterThan", formula1=0,
                             error="El área debe ser un número positivo (m²).", errorTitle="Área inválida")
    ws2.add_data_validation(dv_area)
    dv_area.add(f"C5:C{total_row-1}")

    # Definir un nombre tabla "RecintosNombres" que se usará en la hoja Cargas
    ws2.protection.sheet = False  # no proteger

    # =====================================================================
    # HOJA 3: CARGAS DEDICADAS
    # =====================================================================
    ws3 = wb.create_sheet("Cargas")
    ws3.sheet_view.showGridLines = False
    _title(ws3, 1, 1, "CARGAS DEDICADAS", span_cols=5, size=14)
    ws3.cell(row=2, column=1, value="Lista todos los equipos con circuito propio: cocina eléctrica, calefón, aire, motores, ascensores, etc.").font = Font(name="Calibri", size=10, italic=True, color=GRAY)
    ws3.merge_cells("A2:E2")

    headers_c = ["Nombre del equipo", "Potencia (W)", "Recinto (dropdown)", "Activa (Sí/No)", "Notas"]
    _header_row(ws3, 4, headers_c, col_widths=[32, 14, 22, 14, 28])

    ejemplos_c = [
        ("Cocina eléctrica",         5500, "Cocina",       "Sí", "Solo si NO hay cocina a gas"),
        ("Microondas",               1500, "Cocina",       "Sí", "Encimera de cocina"),
        ("Lavadora",                 2200, "Cocina",       "Sí", "Circuito dedicado 16A"),
        ("Refrigerador",              300, "Cocina",       "Sí", "Comparte enchufes cocina"),
        ("Calefón / termo eléctrico", 5500, "Baño",        "No", "Marcar Sí si NO hay calefón a gas"),
        ("Aire acondicionado",       2200, "Dormitorio 1", "No", "Por cada equipo split"),
    ]
    for i, ej in enumerate(ejemplos_c):
        _data_row(ws3, 5 + i, ej, formats=[None, "#,##0", None, None, None], alt_row=(i % 2 == 1))
    for i in range(len(ejemplos_c), len(ejemplos_c) + 14):
        _data_row(ws3, 5 + i, ["", "", "", "", ""], formats=[None, "#,##0", None, None, None], alt_row=(i % 2 == 1))

    # Total cargas
    total_c_row = 5 + len(ejemplos_c) + 14 + 1
    ws3.cell(row=total_c_row, column=1, value="TOTAL ACTIVAS").font = Font(bold=True, size=11, color=ORANGE_DARK)
    formula = f"=SUMPRODUCT((LOWER(D5:D{total_c_row-1})=\"si\")*B5:B{total_c_row-1})+SUMPRODUCT((LOWER(D5:D{total_c_row-1})=\"sí\")*B5:B{total_c_row-1})"
    ws3.cell(row=total_c_row, column=2, value=formula).font = Font(bold=True, size=11, color=ORANGE_DARK)
    ws3.cell(row=total_c_row, column=2).number_format = "#,##0\" W\""
    for col in range(1, 6):
        c = ws3.cell(row=total_c_row, column=col)
        c.fill = PatternFill("solid", fgColor=ORANGE_LIGHT)
        _set_border(c)

    # Dropdown de recinto enlazado a hoja "Recintos" columna A
    # Usamos rango directo: =Recintos!$A$5:$A$24
    dv_recinto = DataValidation(
        type="list",
        formula1=f"=Recintos!$A$5:$A${5+len(ejemplos)+15-1}",
        allow_blank=True,
    )
    dv_recinto.prompt = "Selecciona el recinto donde está la carga (de la hoja Recintos)"
    dv_recinto.promptTitle = "Ubicación"
    ws3.add_data_validation(dv_recinto)
    dv_recinto.add(f"C5:C{total_c_row-1}")

    # Dropdown Sí/No
    dv_si = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    ws3.add_data_validation(dv_si)
    dv_si.add(f"D5:D{total_c_row-1}")

    # Dropdown potencia (sugerencias)
    dv_pot = DataValidation(type="decimal", operator="greaterThan", formula1=0,
                            error="La potencia debe ser un número positivo (W).", errorTitle="Potencia inválida")
    ws3.add_data_validation(dv_pot)
    dv_pot.add(f"B5:B{total_c_row-1}")

    # Guardar
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def guardar_template_xlsx_a_disco(ruta: Path | str) -> Path:
    """Helper para guardar la plantilla a un archivo en disco."""
    ruta = Path(ruta)
    ruta.write_bytes(generar_template_xlsx())
    return ruta
