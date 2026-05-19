"""Generador de memoria de cálculo en Excel (xlsx).

Construye un workbook con 6 hojas que recogen TODO el cálculo del proyecto
de forma editable: el ingeniero puede revisar fórmulas, modificar supuestos
y recalcular sobre la misma planilla.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.report_sections import seccion_incluida


ORANGE = "C56022"
ORANGE_DARK = "8C3F12"
ORANGE_LIGHT = "F5E6D8"
GRAY = "595959"
BORDER_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=BORDER_THIN, bottom=BORDER_THIN, left=BORDER_THIN, right=BORDER_THIN)

H1 = Font(name="Calibri", size=14, bold=True, color=ORANGE_DARK)
H2 = Font(name="Calibri", size=12, bold=True, color=ORANGE)
H3 = Font(name="Calibri", size=11, bold=True)
LABEL = Font(name="Calibri", size=10, bold=True, color=GRAY)
NORMAL = Font(name="Calibri", size=10)
NUM = Font(name="Calibri", size=10)
FILL_HEAD = PatternFill("solid", fgColor=ORANGE_LIGHT)
FILL_HEAD_DARK = PatternFill("solid", fgColor=ORANGE)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _kv(ws, row, label, value, fmt=None):
    ws.cell(row=row, column=1, value=label).font = LABEL
    c = ws.cell(row=row, column=2, value=value)
    c.font = NORMAL
    if fmt:
        c.number_format = fmt
    return row + 1


def _title(ws, row, text, color_fill=True):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = H2
    if color_fill:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell.fill = PatternFill("solid", fgColor="FAF1E6")
    return row + 2


def _table_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.font = Font(name="Calibri", size=10, bold=True, color=ORANGE_DARK)
        c.fill = FILL_HEAD
        c.alignment = CENTER
        c.border = BORDER
    return row + 1


def _table_row(ws, row, values, start_col=1, formats=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col+i, value=v)
        c.font = NORMAL
        c.border = BORDER
        if formats and i < len(formats) and formats[i]:
            c.number_format = formats[i]
        if isinstance(v, (int, float)):
            c.alignment = RIGHT
    return row + 1


def generar_excel(proyecto: dict, salida: Path | str) -> Path:
    """Genera el workbook XLSX con la memoria de cálculo completa.

    Espera 'proyecto' con las claves: nombre, cliente, tipo_proyecto,
    fecha_creacion, sitio (TERRENO), plano (PLANOS), ric (RIC), fv y eco
    (FV+ECONÓMICO precalculados)."""
    salida = Path(salida)
    wb = Workbook()
    inc = lambda k: seccion_incluida(proyecto, k)

    # === Hoja 1: RESUMEN === (siempre incluida — portada del workbook)
    ws = wb.active
    ws.title = "Resumen"
    _set_widths(ws, [32, 22, 16, 16, 16, 16])
    ws["A1"] = "MEMORIA DE CÁLCULO — SISTEMA SOLAR FOTOVOLTAICO"; ws["A1"].font = H1
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Generado por FV Chile (predimensionamiento técnico-comercial)"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color=GRAY)
    ws.merge_cells("A2:F2")
    r = 4
    r = _title(ws, r, "Proyecto")
    r = _kv(ws, r, "Nombre", proyecto["nombre"])
    r = _kv(ws, r, "Cliente", proyecto.get("cliente",""))
    r = _kv(ws, r, "Tipo de proyecto", proyecto["tipo_proyecto"])
    r = _kv(ws, r, "Fecha de cálculo", proyecto.get("fecha_creacion",""))
    r = _kv(ws, r, "Ubicación", f"{proyecto['sitio']['nombre']}, {proyecto['sitio']['region']}")
    r = _kv(ws, r, "Coordenadas", f"{proyecto['sitio']['lat']}, {proyecto['sitio']['lon']}")
    r = _kv(ws, r, "Altitud (msnm)", proyecto['sitio'].get('altitud_msnm', 0))

    r = _title(ws, r+1, "Cargas RIC")
    ric = proyecto["ric"]
    r = _kv(ws, r, "Superficie total (m²)", ric["area_total_m2"], "#,##0.0")
    r = _kv(ws, r, "Recintos", ric["n_recintos"])
    r = _kv(ws, r, "Carga total instalada (W)", ric["carga_total_instalada_w"], "#,##0")
    r = _kv(ws, r, "Demanda máxima (W)", ric["demanda_maxima_w"], "#,##0")
    r = _kv(ws, r, "Empalme sugerido", ric["tipo_empalme_sugerido"])
    r = _kv(ws, r, "Consumo estimado (kWh/año)", ric["consumo_anual_estimado_kwh"], "#,##0")

    fv = proyecto.get("fv", {})
    eco = proyecto.get("eco", {})
    if fv:
        r = _title(ws, r+1, "Sistema FV propuesto")
        r = _kv(ws, r, "Potencia FV (kWp)", fv["P_kwp"], "#,##0.00")
        r = _kv(ws, r, "N° paneles", fv["N_paneles"])
        r = _kv(ws, r, "Superficie (m²)", fv["superficie_m2"], "#,##0.0")
        r = _kv(ws, r, "Inversor", fv["inversor_modelo"])
        r = _kv(ws, r, "Performance Ratio", fv["PR_anual"], "0.0%")
        r = _kv(ws, r, "Generación anual (kWh)", fv["generacion_anual_kwh"], "#,##0")
        r = _kv(ws, r, "Cobertura solar", fv["cobertura_real"], "0.0%")
    if eco:
        r = _title(ws, r+1, "Análisis económico — 25 años")
        r = _kv(ws, r, "CAPEX total (CLP)", eco["capex_total_clp"], "$#,##0")
        r = _kv(ws, r, "Ahorro año 1 (CLP)", eco["ahorro_total_anual_clp"], "$#,##0")
        r = _kv(ws, r, "Payback simple (años)", eco["payback_simple_anios"], "0.00")
        r = _kv(ws, r, "VAN (CLP)", eco["VAN_clp"], "$#,##0")
        r = _kv(ws, r, "TIR (%)", (eco.get("TIR_pct") or 0)/100, "0.00%")
        r = _kv(ws, r, "LCOE (CLP/kWh)", eco["LCOE_clp_kwh"], "$#,##0")
        r = _kv(ws, r, "CO₂ evitado (tCO₂/año)", eco["CO2_evitado_anual_kg"]/1000, "0.00")

    # === Hoja 2: SITIO === (solo si "sitio" está marcado)
    if inc("sitio"):
        ws2 = wb.create_sheet("Sitio")
        _set_widths(ws2, [22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14])
        ws2["A1"] = "Caracterización del sitio"; ws2["A1"].font = H1
        s = proyecto["sitio"]
        r = 3
        r = _kv(ws2, r, "Ciudad", s["nombre"])
        r = _kv(ws2, r, "Región", s.get("region",""))
        r = _kv(ws2, r, "Latitud (°)", s["lat"])
        r = _kv(ws2, r, "Longitud (°)", s["lon"])
        r = _kv(ws2, r, "Altitud (msnm)", s.get("altitud_msnm",0))
        r += 1
        r = _kv(ws2, r, "Inclinación óptima (°)", s["pvgis"]["slope"])
        r = _kv(ws2, r, "Azimut óptimo (°)", s["pvgis"]["azimuth"])
        r = _kv(ws2, r, "Generación anual por kWp (PVGIS)", s["pvgis"]["E_y"])
        r = _kv(ws2, r, "Irradiación anual (kWh/m²)", s["pvgis"]["H_y"])
        r += 2
        ws2.cell(row=r, column=1, value="DATOS MENSUALES").font = H2
        r += 1
        headers = ["Mes"] + MESES
        r = _table_header(ws2, r, headers)
        r = _table_row(ws2, r, ["E (kWh/kWp) PVGIS"] + s["pvgis"]["monthly_E"], formats=[None]+["0.0"]*12)
        r = _table_row(ws2, r, ["H(i) (kWh/m²) PVGIS"] + s["pvgis"]["monthly_H"], formats=[None]+["0.0"]*12)
        # NASA tiene 11 o 12 valores; rellenar a 12 si falta
        nasa_ghi = list(s["nasa"]["monthly_ghi"]) + [None]*(12-len(s["nasa"]["monthly_ghi"]))
        nasa_t = list(s["nasa"]["monthly_t"]) + [None]*(12-len(s["nasa"]["monthly_t"]))
        nasa_w = list(s["nasa"]["monthly_w"]) + [None]*(12-len(s["nasa"]["monthly_w"]))
        r = _table_row(ws2, r, ["GHI (kWh/m²/día) NASA"] + nasa_ghi, formats=[None]+["0.00"]*12)
        r = _table_row(ws2, r, ["T amb (°C) NASA"] + nasa_t, formats=[None]+["0.0"]*12)
        r = _table_row(ws2, r, ["Viento (m/s) NASA"] + nasa_w, formats=[None]+["0.00"]*12)

    # === Hoja 3: RECINTOS === (sólo si "consumo" está marcado)
    if inc("consumo") and "plano" in proyecto:
        ws3 = wb.create_sheet("Recintos")
        _set_widths(ws3, [8, 30, 16, 14, 14])
        ws3["A1"] = "Recintos detectados del plano"; ws3["A1"].font = H1
        r = 3
        r = _table_header(ws3, r, ["#", "Nombre", "Uso", "Área (m²)", "Perímetro (m)"])
        for rec in proyecto["plano"]["recintos"]:
            r = _table_row(ws3, r, [rec["id"], rec["nombre"], rec["uso"], rec.get("area_m2",0), rec.get("perimetro_m",0)], formats=[None,None,None,"#,##0.0","#,##0.0"])
        r = _table_row(ws3, r, ["", "TOTAL", "", proyecto["plano"]["area_total_m2"], ""], formats=[None,None,None,"#,##0.0",None])
        for col in range(1,6):
            c = ws3.cell(row=r-1, column=col); c.font = Font(name="Calibri", size=10, bold=True); c.fill = FILL_HEAD

    # === Hoja 4: CARGAS RIC === (sólo si "consumo" está marcado)
    if inc("consumo"):
        ws4 = wb.create_sheet("Cargas RIC")
        _set_widths(ws4, [28, 14, 12, 14, 14, 14])
        ws4["A1"] = "Cálculo de cargas según Pliegos RIC SEC"; ws4["A1"].font = H1
        r = 3
        ws4.cell(row=r, column=1, value="Carga por recinto (W/m² × m²)").font = H2; r += 1
        r = _table_header(ws4, r, ["Recinto", "Uso", "Área (m²)", "Alumbrado (W)", "Enchufes (W)", "Subtotal (W)"])
        for c in ric["recintos_carga"]:
            r = _table_row(ws4, r, [c["nombre"], c["uso"], c["area_m2"], c["alumbrado_w"], c["enchufes_w"], c["subtotal_w"]], formats=[None,None,"#,##0.0","#,##0","#,##0","#,##0"])
        r = _table_row(ws4, r, ["Subtotal recintos", "", "", "", "", ric["subtotal_recintos_w"]], formats=[None,None,None,None,None,"#,##0"])
        for col in range(1,7): ws4.cell(row=r-1, column=col).font = Font(bold=True)
        r += 2
        ws4.cell(row=r, column=1, value="Cargas dedicadas").font = H2; r += 1
        r = _table_header(ws4, r, ["Nombre", "", "", "", "Activa", "Potencia (W)"])
        for d in ric["cargas_dedicadas"]:
            r = _table_row(ws4, r, [d["nombre"], "", "", "", "Sí" if d.get("activa") else "No", d["potencia_w"]], formats=[None,None,None,None,None,"#,##0"])
        r = _table_row(ws4, r, ["Subtotal dedicadas", "", "", "", "", ric["subtotal_dedicadas_w"]], formats=[None,None,None,None,None,"#,##0"])
        for col in range(1,7): ws4.cell(row=r-1, column=col).font = Font(bold=True)
        r += 2
        ws4.cell(row=r, column=1, value="Consolidado").font = H2; r += 1
        r = _kv(ws4, r, "Carga total instalada (W)", ric["carga_total_instalada_w"], "#,##0")
        r = _kv(ws4, r, "Factor de demanda", ric["factor_demanda"], "0.00")
        r = _kv(ws4, r, "Carga diversificada (W)", ric["carga_diversificada_w"], "#,##0")
        r = _kv(ws4, r, "Factor de simultaneidad", ric["factor_simultaneidad"], "0.00")
        r = _kv(ws4, r, "DEMANDA MÁXIMA (W)", ric["demanda_maxima_w"], "#,##0")
        r = _kv(ws4, r, "Corriente nominal (A)", ric["corriente_nominal_a"], "0.0")
        r = _kv(ws4, r, "Empalme sugerido", ric["tipo_empalme_sugerido"])
        r = _kv(ws4, r, "Conexión", ric["conexion"])
        r = _kv(ws4, r, "Factor de potencia asumido", ric["factor_potencia"], "0.00")

    # === Hoja 5: DISEÑO FV === (sólo si "fv_dimensionamiento" está marcado)
    if fv and inc("fv_dimensionamiento"):
        ws5 = wb.create_sheet("Diseño FV")
        _set_widths(ws5, [28, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14])
        ws5["A1"] = "Dimensionamiento del sistema fotovoltaico"; ws5["A1"].font = H1
        r = 3
        r = _kv(ws5, r, "Potencia FV (kWp)", fv["P_kwp"], "#,##0.00")
        r = _kv(ws5, r, "Paneles", f"{fv['N_paneles']} × 550 Wp")
        r = _kv(ws5, r, "Superficie (m²)", fv["superficie_m2"], "#,##0.0")
        r = _kv(ws5, r, "Inversor", fv["inversor_modelo"])
        r = _kv(ws5, r, "P AC inversor (kW)", fv["inversor_P_AC_kw"], "#,##0.0")
        r = _kv(ws5, r, "T celda promedio (°C)", fv["T_celda_promedio_c"], "0.0")
        r = _kv(ws5, r, "Performance Ratio", fv["PR_anual"], "0.0%")
        r = _kv(ws5, r, "Factor de planta", fv["factor_planta"], "0.0%")
        r += 2
        ws5.cell(row=r, column=1, value="Desglose de pérdidas").font = H2; r += 1
        r = _table_header(ws5, r, ["Pérdida", "%"])
        for pp in fv["perdidas"]:
            r = _table_row(ws5, r, [pp["nombre"], pp["pct"]], formats=[None,"0.00"])
        r += 1
        ws5.cell(row=r, column=1, value="Generación mensual").font = H2; r += 1
        r = _table_header(ws5, r, ["Mes"] + MESES)
        r = _table_row(ws5, r, ["Generación (kWh)"] + fv["generacion_mensual_kwh"], formats=[None]+["#,##0"]*12)
        r += 1
        r = _kv(ws5, r, "Generación anual (kWh)", fv["generacion_anual_kwh"], "#,##0")
        r = _kv(ws5, r, "Cobertura real", fv["cobertura_real"], "0.0%")
        r = _kv(ws5, r, "Autoconsumo (kWh/año)", fv["autoconsumo_kwh"], "#,##0")
        r = _kv(ws5, r, "Inyección a red (kWh/año)", fv["inyeccion_kwh"], "#,##0")
        r = _kv(ws5, r, "Compra a red (kWh/año)", fv["compra_red_kwh"], "#,##0")

    # === Hoja 6: ECONÓMICO === (sólo si alguna sección económica está marcada)
    if eco and (inc("capex") or inc("metricas_economicas") or inc("flujo_caja")):
        ws6 = wb.create_sheet("Económico")
        _set_widths(ws6, [22, 18, 18, 18])
        ws6["A1"] = "Análisis económico — 25 años"; ws6["A1"].font = H1
        r = 3
        ws6.cell(row=r, column=1, value="Desglose CAPEX (CLP)").font = H2; r += 1
        r = _table_header(ws6, r, ["Partida", "CLP", "%"])
        for k, v in eco["capex_desglose"].items():
            if v:
                r = _table_row(ws6, r, [k, v, v/eco["capex_total_clp"]], formats=[None,"$#,##0","0.0%"])
        r = _table_row(ws6, r, ["TOTAL", eco["capex_total_clp"], 1], formats=[None,"$#,##0","0.0%"])
        for col in range(1,4): ws6.cell(row=r-1, column=col).font = Font(bold=True)
        r += 2
        ws6.cell(row=r, column=1, value="Métricas").font = H2; r += 1
        r = _kv(ws6, r, "Ahorro anual y1 (CLP)", eco["ahorro_total_anual_clp"], "$#,##0")
        r = _kv(ws6, r, "OPEX anual (CLP)", eco["opex_anual_clp"], "$#,##0")
        r = _kv(ws6, r, "Payback simple (años)", eco["payback_simple_anios"], "0.00")
        r = _kv(ws6, r, "VAN (CLP)", eco["VAN_clp"], "$#,##0")
        r = _kv(ws6, r, "TIR (%)", (eco.get("TIR_pct") or 0)/100, "0.00%")
        r = _kv(ws6, r, "LCOE (CLP/kWh)", eco["LCOE_clp_kwh"], "$#,##0")
        r = _kv(ws6, r, "Tasa de descuento", eco["tasa_descuento"], "0.0%")
        r += 2
        ws6.cell(row=r, column=1, value="Flujo de caja anual").font = H2; r += 1
        r = _table_header(ws6, r, ["Año", "Flujo (CLP)", "Acumulado (CLP)"])
        for y in range(len(eco["flujo_caja_anual_clp"])):
            r = _table_row(ws6, r, [y, eco["flujo_caja_anual_clp"][y], eco["flujo_acumulado_clp"][y]], formats=[None,"$#,##0","$#,##0"])

    wb.save(salida)
    return salida
